import os
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import anthropic
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pydantic import BaseModel, Field

# override=True ensures the .env file is authoritative even if a stray
# (possibly empty) variable of the same name already exists in the environment.
# On Streamlit Cloud there is no .env file, so this overrides nothing and the
# secrets-bridged values survive.
load_dotenv(override=True)
API_KEY = os.getenv("GOOGLE_PLACES_API_KEY")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")

TEXT_SEARCH_URL = "https://maps.googleapis.com/maps/api/place/textsearch/json"
DETAILS_URL = "https://maps.googleapis.com/maps/api/place/details/json"
DETAILS_FIELDS = "website,formatted_phone_number"
DEFAULT_MAX_RESULTS = 25

QUALIFICATION_MODEL = "claude-haiku-4-5"

EMAIL_RE = re.compile(r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b")
GENERIC_EMAIL_PREFIXES = {
    "info", "contact", "hello", "hi", "support", "sales", "admin", "office",
    "help", "team", "mail", "service", "reception", "reservations", "bookings",
    "inquiries", "enquiries", "questions", "customerservice",
}
# Homepage first, then common contact-page slugs across languages.
# /kontakt covers Polish & German; /contacto Spanish; /contatti Italian;
# /contact-us & /contact English. We only try later paths if the homepage
# yields no emails (see early-exit in scrape_emails), so extra paths add no
# cost for sites that already expose an email up front.
CONTACT_PATHS = [
    "/", "/contact", "/contact-us", "/kontakt", "/contacto", "/contatti",
]
SKIP_DOMAINS_FOR_SCRAPING = (
    "facebook.com", "instagram.com", "yelp.com", "tripadvisor.",
    "google.com", "linkedin.com", "twitter.com", "x.com",
)
SCRAPER_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)
SCRAPER_TIMEOUT = 8


class LeadQualification(BaseModel):
    score: int = Field(ge=1, le=10, description="Lead quality on a 1-10 scale.")
    reasoning: str = Field(
        description="One or two sentences explaining the score, grounded in the data."
    )
    possible_chain: bool = Field(
        description=(
            "True if the business name clearly matches a known national or regional chain "
            "(e.g., Starbucks, Dunkin', Aspen Dental, Great Clips, Walgreens). "
            "False if the name appears independent or you are unsure."
        )
    )


QUALIFICATION_SYSTEM = """You are a lead-qualification assistant for a generic small-to-medium business (SMB) outreach workflow.
Given publicly available data about a local business, rate how promising it is as a cold-outreach prospect.

BE CALIBRATED. Most decent businesses fall in the 5-7 range. A 9 or 10 should be RARE — reserve those scores for businesses that are exceptional on multiple dimensions. If you find yourself giving every lead an 8 or 9, you are not differentiating enough.

Score 1-10 using these bands:
- 9-10: exceptional — 1000+ reviews at 4.6+, polished independent website, no concerns whatsoever
- 7-8:  strong lead — solid review volume (200-1000), good rating (4.3+), real website, minor or no concerns
- 5-6:  decent lead — moderate signals across the board, or strong in some dimensions and weak in others
- 3-4:  weak — meaningful concerns (e.g., thin reviews, no website, mediocre rating)
- 1-2:  poor — multiple red flags, OR a clear category mismatch (see override below)

Signals that RAISE the score:
- Real independent website on its own domain (not Facebook, Yelp, or a generic platform page)
- High review volume AND high rating together (200+ reviews at 4.3+)
- Appears independent / locally owned (not an obvious national chain)

Signals that LOWER the score:
- No website at all
- No phone number on file
- Rating below 3.5
- Very few reviews (under 10) — too little signal to trust
- A perfect 5.0 rating with fewer than 5 reviews (classic fake-review pattern)
- Obvious national or regional chain

IMPORTANT CALIBRATION RULES:
1. Matching the search category is the BASELINE expectation, not a bonus. Do NOT raise scores just because the business is "on topic" — that's assumed.
2. Do NOT pile on positive language for routine-good leads. A business with 500 reviews at 4.5 and a website is a 7 or 8, not a 9.
3. Reserve 9-10 for leads that genuinely stand out. If a batch of 10 leads all look comparable, most should be 6-8.

CATEGORY MISMATCH OVERRIDE:
If the business is clearly NOT what the user searched for (e.g., a chiropractor returned for "marijuana dispensaries", a hardware store for "restaurants"), score it 1-2 regardless of other signals and explain the mismatch in the reasoning. Look for clues in the business name, categories field, and website domain. When data is ambiguous, do not invent a mismatch — only flag clear ones.

Keep the reasoning to one or two sentences. Be specific — cite the actual rating, review count, or website. Do not invent information."""


def compute_objective_flags(row):
    flags = []
    website = row["Website"]
    has_real_website = (
        website != "N/A"
        and not any(d in website.lower() for d in SKIP_DOMAINS_FOR_SCRAPING)
    )
    if not has_real_website:
        flags.append("no_website")
    if row["Phone"] == "N/A":
        flags.append("no_phone")
    reviews = row["Number of Reviews"]
    rating = row["Rating"]
    reviews_num = reviews if isinstance(reviews, (int, float)) else None
    rating_num = rating if isinstance(rating, (int, float)) else None
    if reviews_num is not None and reviews_num < 10:
        flags.append("thin_reviews")
    if rating_num is not None and rating_num < 3.5:
        flags.append("low_rating")
    if reviews_num is not None and rating_num is not None:
        if reviews_num >= 500 and rating_num >= 4.5:
            flags.append("established")
    return flags


def classify_email(email):
    local = email.split("@", 1)[0].lower()
    return "generic" if local in GENERIC_EMAIL_PREFIXES else "personal"


def _looks_like_real_email(email):
    lower = email.lower()
    # Regex sometimes catches things like logo@2x.png (retina image refs),
    # Google Calendar feed IDs, or HTML form placeholder text — filter those out.
    if any(lower.endswith(ext) for ext in (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp")):
        return False
    junk_substrings = (
        "@sentry", "@example.", "@example.com", "@wixpress", "@2x.", "@3x.", "@domain.",
        ".calendar.google.com",                    # embedded calendar feed IDs
        "@email.com",                              # placeholder: name@email.com, your@email.com
        "@yourdomain", "@yoursite", "@yourcompany",
        "@test.com", "@mail.com",
        "u003",                                    # JS Unicode-escape leakage (e.g., >info@...)
    )
    return not any(j in lower for j in junk_substrings)


def scrape_emails(website_url):
    if not website_url or website_url == "N/A":
        return []
    lower = website_url.lower()
    if any(d in lower for d in SKIP_DOMAINS_FOR_SCRAPING):
        return []
    parsed = urlparse(website_url)
    if not parsed.scheme or not parsed.netloc:
        return []
    base = f"{parsed.scheme}://{parsed.netloc}"

    headers = {"User-Agent": SCRAPER_USER_AGENT}
    found = set()
    for path in CONTACT_PATHS:
        try:
            resp = requests.get(
                urljoin(base, path),
                headers=headers,
                timeout=SCRAPER_TIMEOUT,
                allow_redirects=True,
            )
        except requests.RequestException:
            continue
        if resp.status_code != 200:
            continue
        if "text/html" not in resp.headers.get("Content-Type", "").lower():
            continue
        for match in EMAIL_RE.findall(resp.text):
            if _looks_like_real_email(match):
                found.add(match.lower())
        # If we already found some emails, don't keep hitting more pages.
        if found:
            break

    return [(email, classify_email(email)) for email in sorted(found)]


def build_qualification_prompt(row, query):
    return (
        f"Original search query: {query}\n\n"
        "Evaluate this business as a cold-outreach prospect.\n\n"
        f"Business name:  {row['Business Name']}\n"
        f"Categories:     {row['Categories']}\n"
        f"Address:        {row['Address']}\n"
        f"Website:        {row['Website']}\n"
        f"Phone:          {row['Phone']}\n"
        f"Rating:         {row['Rating']}\n"
        f"Review count:   {row['Number of Reviews']}"
    )


def qualify_lead(row, client, query):
    response = client.messages.parse(
        model=QUALIFICATION_MODEL,
        max_tokens=1000,
        system=QUALIFICATION_SYSTEM,
        messages=[{"role": "user", "content": build_qualification_prompt(row, query)}],
        output_format=LeadQualification,
    )
    return response.parsed_output


def _text_search_request(params, attempt=0):
    response = requests.get(TEXT_SEARCH_URL, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()
    status = data.get("status")
    if status in ("OK", "ZERO_RESULTS"):
        return data
    # A pagetoken can take several seconds to become valid on Google's side.
    # INVALID_REQUEST on a pagetoken call almost always means "not ready yet" — retry.
    if status == "INVALID_REQUEST" and "pagetoken" in params and attempt < 4:
        time.sleep(2 + attempt * 2)
        return _text_search_request(params, attempt + 1)
    raise RuntimeError(
        f"Google Places error: {status} — {data.get('error_message', '')}"
    )


def search_places(query, max_results):
    results = []
    params = {"query": query, "key": API_KEY}
    is_first_page = True
    while len(results) < max_results:
        try:
            data = _text_search_request(params)
        except RuntimeError as e:
            if is_first_page:
                raise
            print(f"  (skipping further pages — {e}. Keeping {len(results)} results.)")
            break
        results.extend(data.get("results", []))
        next_token = data.get("next_page_token")
        if not next_token or len(results) >= max_results:
            break
        is_first_page = False
        time.sleep(3)
        params = {"pagetoken": next_token, "key": API_KEY}
    return results[:max_results]


def fetch_details(place_id):
    # A transient network error on a single lead's details call should not
    # crash the whole run — return empty so the lead keeps its search data
    # (name, address, rating) with phone/website falling back to "N/A".
    params = {"place_id": place_id, "fields": DETAILS_FIELDS, "key": API_KEY}
    for attempt in range(3):
        try:
            response = requests.get(DETAILS_URL, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            if data.get("status") != "OK":
                return {}
            return data.get("result", {})
        except requests.RequestException:
            if attempt < 2:
                time.sleep(1 + attempt)
                continue
            return {}
    return {}


def build_row(place):
    details = fetch_details(place["place_id"])
    types = place.get("types") or []
    return {
        "place_id": place["place_id"],
        "Business Name": place.get("name") or "N/A",
        "Address": place.get("formatted_address") or "N/A",
        "Phone": details.get("formatted_phone_number") or "N/A",
        "Website": details.get("website") or "N/A",
        "Rating": place.get("rating", "N/A"),
        "Number of Reviews": place.get("user_ratings_total", "N/A"),
        "Categories": ", ".join(types) if types else "N/A",
    }


def slugify(text):
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text[:60] or "leads"


def get_headers(include_qualification=True, include_enrichment=False):
    base_headers = [
        "Business Name", "Address", "Phone", "Website",
        "Rating", "Number of Reviews", "Categories",
    ]
    headers = []
    if include_qualification:
        headers.append("Score")
    headers.extend(base_headers)
    if include_qualification:
        headers.extend(["Reasoning", "Flags"])
    if include_enrichment:
        headers.extend(["Emails (Generic)", "Emails (Personal)"])
    return headers


def build_workbook(rows, include_qualification=True, include_enrichment=False):
    headers = get_headers(include_qualification, include_enrichment)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row in rows:
        ws.append([row.get(h, "N/A") for h in headers])

    for col_idx, header in enumerate(headers, start=1):
        longest = max(
            [len(str(header))]
            + [len(str(row.get(header, ""))) for row in rows]
        )
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = min(longest + 2, 60)

    ws.freeze_panes = "A2"
    return wb


def write_excel(rows, query, include_qualification=True, include_enrichment=False):
    wb = build_workbook(rows, include_qualification, include_enrichment)
    filename = f"{slugify(query)}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    output_path = Path(filename)
    wb.save(output_path)
    return output_path


def run_pipeline(query, max_results, include_qualification, include_enrichment, log=None):
    """Run the full pipeline. Returns the list of lead rows.

    log(message) is called with progress strings if provided.
    """
    def emit(msg):
        if log:
            log(msg)

    emit(f"Searching Google Places for: {query}")
    places = search_places(query, max_results)
    if not places:
        return []

    emit(f"Found {len(places)}. Fetching contact details for each...")
    rows = []
    for i, place in enumerate(places, start=1):
        rows.append(build_row(place))
        emit(f"  [{i}/{len(places)}] {place.get('name', '(unnamed)')}")

    if include_qualification:
        emit("Qualifying leads with Claude...")
        claude = anthropic.Anthropic()
        for i, row in enumerate(rows, start=1):
            try:
                result = qualify_lead(row, claude, query)
                flags = compute_objective_flags(row)
                if result.possible_chain:
                    flags.append("possible_chain")
                row["Score"] = result.score
                row["Reasoning"] = result.reasoning
                row["Flags"] = ", ".join(flags) if flags else "none"
                emit(f"  [{i}/{len(rows)}] {row['Business Name'][:40]:40s}  score={result.score}")
            except Exception as e:
                row["Score"] = "N/A"
                row["Reasoning"] = f"Qualification failed: {e}"
                row["Flags"] = "N/A"
                emit(f"  [{i}/{len(rows)}] {row['Business Name'][:40]:40s}  FAILED ({e})")
        rows.sort(key=lambda r: r["Score"] if isinstance(r["Score"], int) else -1, reverse=True)

    if include_enrichment:
        emit("Scraping websites for contact emails...")
        for i, row in enumerate(rows, start=1):
            results = scrape_emails(row["Website"])
            generic = [e for e, t in results if t == "generic"]
            personal = [e for e, t in results if t == "personal"]
            row["Emails (Generic)"] = ", ".join(generic) if generic else "N/A"
            row["Emails (Personal)"] = ", ".join(personal) if personal else "N/A"
            found_count = len(generic) + len(personal)
            emit(f"  [{i}/{len(rows)}] {row['Business Name'][:40]:40s}  emails={found_count}")

    return rows


def main():
    if not API_KEY:
        print("ERROR: GOOGLE_PLACES_API_KEY is missing.")
        print("Create a file named .env in this folder with one line:")
        print("GOOGLE_PLACES_API_KEY=your_key_here")
        return

    query = input(
        "What kind of businesses are you searching for?\n"
        "(e.g., 'dental offices in Milwaukee, Wisconsin')\n> "
    ).strip()
    if not query:
        print("No query entered. Exiting.")
        return

    raw_max = input(
        f"How many results? (press Enter for {DEFAULT_MAX_RESULTS}) > "
    ).strip()
    try:
        max_results = int(raw_max) if raw_max else DEFAULT_MAX_RESULTS
    except ValueError:
        print(f"Not a number — using {DEFAULT_MAX_RESULTS}.")
        max_results = DEFAULT_MAX_RESULTS

    include_qualification = False
    if ANTHROPIC_API_KEY:
        raw = input("Include AI qualification (score, reasoning, flags)? [Y/n] > ").strip().lower()
        include_qualification = raw in ("", "y", "yes")
    else:
        print("(AI qualification unavailable — ANTHROPIC_API_KEY not set in .env)")

    raw = input("Include email enrichment (scrape business websites for emails)? [Y/n] > ").strip().lower()
    include_enrichment = raw in ("", "y", "yes")

    rows = run_pipeline(
        query=query,
        max_results=max_results,
        include_qualification=include_qualification,
        include_enrichment=include_enrichment,
        log=print,
    )
    if not rows:
        print("No businesses found. Try a different query.")
        return

    output_path = write_excel(
        rows, query,
        include_qualification=include_qualification,
        include_enrichment=include_enrichment,
    )
    print(f"\nDone. Saved {len(rows)} leads to:\n  {output_path.resolve()}")


if __name__ == "__main__":
    main()
